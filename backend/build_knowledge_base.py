"""
Pulls alarm/fault data from GitHub repos and builds a unified knowledge base JSON.
Run once to populate data/knowledge_base.json, or on a schedule to refresh.

Sources (in priority order):
  1. Alarms.xlsx  — Theresa's masterlist (local file or fetched from GitHub repo)
  2. generacclean/generac-home-error-catalog  (self_test, configuration, device_alarm YAMLs)
  3. neurio/pwrinverter  (ExtendedDescriptions/*.md)
  4. neurio/pwrbmu      (ExtendedDescriptions/*.md)
  5. neurio/reef        (headend-events.json — GMS/PLCHE/Manta alarms)

Automation: A Power Automate flow pushes Alarms.xlsx to BenyaminJalali/alarm repo
whenever Theresa saves. The container reads from data/Alarms.xlsx at startup.
"""

import os, json, base64, re, subprocess, sys
from pathlib import Path

GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "")
OUT_PATH = Path(__file__).parent.parent / "data" / "knowledge_base.json"
XLSX_PATH = Path(__file__).parent.parent / "data" / "Alarms.xlsx"
SUPPLEMENTAL_PATH = Path(__file__).parent.parent / "data" / "supplemental_kb.json"

# ── helpers ───────────────────────────────────────────────────────────────────

def gh_api(path: str) -> dict | list | None:
    env = {**os.environ, "GITHUB_TOKEN": GITHUB_TOKEN}
    result = subprocess.run(
        ["gh", "api", path],
        capture_output=True, text=True, env=env
    )
    if result.returncode != 0:
        print(f"  WARN: gh api {path} failed: {result.stderr[:100]}", file=sys.stderr)
        return None
    return json.loads(result.stdout)

def gh_file_content(repo: str, file_path: str) -> str | None:
    data = gh_api(f"repos/{repo}/contents/{file_path}")
    if not data or "content" not in data:
        return None
    return base64.b64decode(data["content"]).decode("utf-8", errors="replace")

def gh_file_bytes(repo: str, file_path: str) -> bytes | None:
    data = gh_api(f"repos/{repo}/contents/{file_path}")
    if not data or "content" not in data:
        return None
    return base64.b64decode(data["content"])

def gh_tree(repo: str) -> list[str]:
    data = gh_api(f"repos/{repo}/git/trees/HEAD?recursive=1")
    if not data:
        return []
    return [item["path"] for item in data.get("tree", []) if item["type"] == "blob"]


# ── masterlist from Excel ─────────────────────────────────────────────────────

def load_masterlist_from_xlsx() -> list[dict]:
    """
    Load alarm entries from Alarms.xlsx — Theresa's masterlist.
    Tries GitHub repo first (Power Automate keeps it current), falls back to local file.
    Reads both 'Alarms Masterlist' sheet and 'MI List' sheet (Micro Inverter alarms).
    """
    xlsx_path = XLSX_PATH

    # Try to fetch latest from GitHub repo first (Power Automate keeps this current)
    print("  Checking GitHub repo for latest Alarms.xlsx...")
    xlsx_bytes = gh_file_bytes("BenyaminJalali/alarm", "data/Alarms.xlsx")
    if xlsx_bytes:
        xlsx_path = Path("/tmp/Alarms_latest.xlsx")
        xlsx_path.write_bytes(xlsx_bytes)
        print("  Using latest Alarms.xlsx from GitHub repo")
    elif XLSX_PATH.exists():
        print("  Using local Alarms.xlsx")
    else:
        print("  WARN: No Alarms.xlsx found — skipping masterlist", file=sys.stderr)
        return []

    try:
        import openpyxl
    except ImportError:
        print("  WARN: openpyxl not installed — skipping masterlist", file=sys.stderr)
        return []

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    all_entries = []

    # ── Sheet 1: Alarms Masterlist (INV, BMU, GMS, PLCHE, MAN) ──
    if "Alarms Masterlist" in wb.sheetnames:
        ws = wb["Alarms Masterlist"]
        rows = list(ws.iter_rows(values_only=True))
        header = [str(h).strip() if h else "" for h in rows[0]]

        def col(name, h=header):
            try: return h.index(name)
            except ValueError: return None

        idx = {
            "device":        col("Device"),
            "alarm_name":    col("Firmware raw alarm"),
            "friendly":      col("Alarm Name"),
            "alarm_code":    col("Alarm code"),
            "description":   col("Description"),
            "corrective":    col("Corrective Action"),
            "internal":      col("Internal Notes"),
            "severity":      col("Severity"),
            "support_vis":   col("Support Visibility"),
            "installer_vis": col("Installer Visibility"),
            "homeowner_vis": col("Homeowner Visibility"),
            "alarm_type":    col("Alarm Type"),
        }

        def get(row, key, i=idx):
            ci = i.get(key)
            if ci is None or ci >= len(row): return None
            v = row[ci]
            if v is None: return None
            if isinstance(v, bool): return v
            return str(v).strip()

        for row in rows[1:]:
            device = get(row, "device")
            alarm_name = get(row, "alarm_name")
            if not device or not alarm_name:
                continue
            all_entries.append({
                "device": device, "alarm_name": alarm_name.strip(),
                "friendly_name": get(row, "friendly") or "",
                "alarm_code": get(row, "alarm_code") or "",
                "description": get(row, "description") or "",
                "corrective_action": get(row, "corrective") or "",
                "internal_notes": get(row, "internal") or "",
                "severity": get(row, "severity") or "",
                "alarm_type": get(row, "alarm_type") or "",
                "visible_support": bool(get(row, "support_vis")),
                "visible_installer": bool(get(row, "installer_vis")),
                "visible_homeowner": bool(get(row, "homeowner_vis")),
            })
        print(f"  Alarms Masterlist: {len(all_entries)} entries")
    else:
        print("  WARN: 'Alarms Masterlist' sheet not found", file=sys.stderr)

    # ── Sheet 2: MI List (Micro Inverter alarms) ──
    mi_count = 0
    if "MI List" in wb.sheetnames:
        ws = wb["MI List"]
        rows = list(ws.iter_rows(values_only=True))
        header = [str(h).strip() if h else "" for h in rows[0]]

        def col_mi(name, h=header):
            try: return h.index(name)
            except ValueError: return None

        idx_mi = {
            "device":        col_mi("Device"),
            "alarm_name":    col_mi("Firmware raw alarm"),
            "friendly":      col_mi("Alarm Name"),
            "alarm_code":    col_mi("Alarm code"),
            "description":   col_mi("Description"),
            "corrective":    col_mi("Corrective Action"),
            "internal":      col_mi("Internal Notes"),
            "priority":      col_mi("Priority (1 high, 3 low, 0 as unlikely to show)"),
            "support_vis":   col_mi("Support Visibility"),
            "installer_vis": col_mi("Installer Visibility"),
            "alarm_type":    col_mi("Alarm Type"),
        }

        def get_mi(row, key, i=idx_mi):
            ci = i.get(key)
            if ci is None or ci >= len(row): return None
            v = row[ci]
            if v is None: return None
            if isinstance(v, bool): return v
            return str(v).strip()

        priority_map = {"1": "Critical (Severity 1)", "2": "High (Severity 2)", "3": "Medium (Severity 3)"}

        for row in rows[1:]:
            device = get_mi(row, "device")
            alarm_name = get_mi(row, "alarm_name")
            if not device or not alarm_name:
                continue
            priority = get_mi(row, "priority") or "3"
            severity = priority_map.get(str(priority).split(".")[0], "Medium (Severity 3)")
            all_entries.append({
                "device": device, "alarm_name": alarm_name.strip(),
                "friendly_name": get_mi(row, "friendly") or "",
                "alarm_code": get_mi(row, "alarm_code") or "",
                "description": get_mi(row, "description") or "",
                "corrective_action": get_mi(row, "corrective") or "",
                "internal_notes": get_mi(row, "internal") or "",
                "severity": severity,
                "alarm_type": get_mi(row, "alarm_type") or "",
                "visible_support": bool(get_mi(row, "support_vis")),
                "visible_installer": bool(get_mi(row, "installer_vis")),
                "visible_homeowner": False,
            })
            mi_count += 1
        print(f"  MI List: {mi_count} entries")
    else:
        print("  WARN: 'MI List' sheet not found", file=sys.stderr)

    print(f"  Total from Excel: {len(all_entries)} entries")
    return all_entries


# ── helpers ───────────────────────────────────────────────────────────────────

def infer_device_from_id(alarm_id: str) -> str:
    """Infer device tag from alarm id like 'configuration.battery.some_alarm'."""
    s = alarm_id.lower()
    segments = s.split(".")
    # Check second segment (e.g. configuration.battery.*)
    device_seg = segments[1] if len(segments) > 1 else ""
    if device_seg in ("battery",):
        return "BMU"
    if device_seg in ("inverter",):
        return "INV"
    if device_seg in ("gateway",):
        return "GMS"
    if device_seg in ("disconnect_switch", "sds", "smart_disconnect", "manta"):
        return "MANTA"
    # Fallback: check anywhere in the id
    if "battery" in s or ".bmu" in s:
        return "BMU"
    if "inverter" in s or ".inv" in s:
        return "INV"
    if "gateway" in s or ".gms" in s:
        return "GMS"
    if "disconnect_switch" in s or ".sds." in s or "manta" in s:
        return "MANTA"
    return ""


# ── parsers ───────────────────────────────────────────────────────────────────

def parse_yaml_entry(content: str, file_path: str) -> dict | None:
    entry = {"source_file": file_path, "id": "", "engineering": {}, "product": {}}
    section = None
    eng_ts, prod_ts, int_ts = [], [], []
    current_step_list = None
    buffer = []

    def flush_buffer(target: dict, key: str):
        if buffer:
            target[key] = " ".join(" ".join(buffer).split())
            buffer.clear()

    for raw_line in content.splitlines():
        line = raw_line.rstrip()
        stripped = line.lstrip()

        if stripped.startswith("#"):
            continue

        if line == "engineering:":
            section = "engineering"
            continue
        if line == "product:":
            section = "product"
            continue
        if stripped.startswith("troubleshooting:") and section == "engineering":
            current_step_list = eng_ts
            continue
        if stripped.startswith("internal_troubleshooting:") and section == "product":
            current_step_list = int_ts
            continue
        if stripped.startswith("troubleshooting:") and section == "product":
            current_step_list = prod_ts
            continue

        m = re.match(r'^id:\s*(.+)', line)
        if m:
            entry["id"] = m.group(1).strip()
            continue

        if current_step_list is not None:
            m = re.match(r'\s*-\s*step:\s*(.*)', line)
            if m:
                current_step_list.append(m.group(1).strip())
                continue
            if stripped and not re.match(r'\w+:', stripped):
                if current_step_list:
                    current_step_list[-1] += " " + stripped
                continue

        m = re.match(r'\s{2}(\w+):\s*(.*)', line)
        if m and section:
            key, val = m.group(1), m.group(2).strip()
            val = val.strip('"\'')
            current_step_list = None
            target = entry[section] if section in entry else {}
            if val in ("null", "~", ""):
                target[key] = None
            elif val in ("true", "True"):
                target[key] = True
            elif val in ("false", "False"):
                target[key] = False
            elif val.startswith(">"):
                buffer.clear()
                entry["_buffer_target"] = (section, key)
            else:
                target[key] = val
                entry[section] = target
            continue

        if "_buffer_target" in entry and stripped:
            buffer.append(stripped)
        elif "_buffer_target" in entry and not stripped:
            sec, key = entry.pop("_buffer_target")
            entry[sec][key] = " ".join(" ".join(buffer).split())
            buffer.clear()

    if "_buffer_target" in entry:
        sec, key = entry.pop("_buffer_target")
        entry[sec][key] = " ".join(" ".join(buffer).split())
        buffer.clear()

    if eng_ts:
        entry["engineering"]["troubleshooting"] = eng_ts
    if prod_ts:
        entry["product"]["troubleshooting"] = prod_ts
    if int_ts:
        entry["product"]["internal_troubleshooting"] = int_ts

    return entry if entry.get("id") or entry.get("engineering") else None


def parse_extended_md(content: str, alarm_name: str) -> dict:
    sections = {"Hardware": "", "Trigger": "", "Thresholds": "", "Troubleshooting": []}
    current = None
    steps = []
    text_buf = []

    for line in content.splitlines():
        stripped = line.strip()
        if stripped.startswith("## "):
            if current == "Troubleshooting":
                sections["Troubleshooting"] = steps
            elif current and text_buf:
                sections[current] = " ".join(text_buf).strip()
            current = stripped[3:].strip()
            text_buf = []
            steps = []
        elif current == "Troubleshooting":
            m = re.match(r'\d+\.\s+(.*)', stripped)
            if m:
                steps.append(m.group(1))
            elif stripped and not stripped.startswith("#"):
                steps.append(stripped)
        elif current:
            if stripped:
                text_buf.append(stripped)

    if current == "Troubleshooting":
        sections["Troubleshooting"] = steps
    elif current and text_buf:
        sections[current] = " ".join(text_buf).strip()

    return {
        "id": f"device_alarm.fw.{alarm_name.lower()}",
        "alarm_name": alarm_name,
        "source": "firmware_extended_desc",
        "engineering": {
            "trigger": sections.get("Trigger", ""),
            "threshold": sections.get("Thresholds", ""),
            "hardware": sections.get("Hardware", ""),
            "troubleshooting": sections.get("Troubleshooting", []),
        }
    }


def parse_headend_events(content: str) -> list[dict]:
    try:
        data = json.loads(content)
    except Exception:
        return []
    entries = []
    for ev in data.get("Events", []):
        alarm = ev.get("Alarm", {})
        cond = ev.get("Condition", {})
        entries.append({
            "id": f"device_alarm.gms.{alarm.get('Name', '').lower()}",
            "alarm_name": alarm.get("Name", ""),
            "device": "GMS/MANTA",
            "source": "headend_events",
            "engineering": {
                "description": alarm.get("Description", ""),
                "trigger": ev.get("Failure", ""),
                "alarm_type": alarm.get("Type", ""),
                "alarm_id": alarm.get("ID", ""),
                "condition_type": cond.get("Type", ""),
                "threshold": cond.get("Threshold"),
                "activation_time_s": cond.get("Activation_time_s"),
            }
        })
    return entries


# ── main build ────────────────────────────────────────────────────────────────

def build():
    kb = {"version": "1.0", "sources": [], "entries": []}
    entry_map: dict[str, dict] = {}

    # 1. Seed from Theresa's masterlist (Alarms.xlsx)
    print("Loading masterlist from Alarms.xlsx...")
    masterlist = load_masterlist_from_xlsx()
    for row in masterlist:
        key = f"{row['device']}:{row['alarm_name']}"
        entry_map[key] = {
            "id": f"device_alarm.{row['device'].lower()}.{row['alarm_name'].lower()}",
            "alarm_name": row["alarm_name"],
            "device": row["device"],
            "friendly_name": row["friendly_name"],
            "alarm_code": row["alarm_code"],
            "severity": row["severity"],
            "alarm_type": row["alarm_type"],
            "visibility": {
                "support": row["visible_support"],
                "installer": row["visible_installer"],
                "homeowner": row["visible_homeowner"],
            },
            "engineering": {
                "description": row["description"],
                "internal_notes": row["internal_notes"],
            },
            "product": {
                "corrective_action": row["corrective_action"],
            },
            "sources": ["masterlist"],
        }
    print(f"  {len(entry_map)} masterlist entries loaded")
    kb["sources"].append("masterlist")

    # 2. Pull catalog YAMLs
    print("Pulling error catalog YAMLs...")
    all_paths = gh_tree("generacclean/generac-home-error-catalog")
    catalog_paths = [p for p in all_paths if p.startswith("catalog/") and p.endswith(".yaml")]
    print(f"  Found {len(catalog_paths)} catalog YAML files")
    for path in catalog_paths:
        content = gh_file_content("generacclean/generac-home-error-catalog", path)
        if not content:
            continue
        parsed = parse_yaml_entry(content, path)
        if parsed:
            cid = parsed.get("id", path)
            parsed["sources"] = ["error_catalog"]
            # Infer device from alarm id (e.g. configuration.battery.* → BMU)
            if not parsed.get("device"):
                inferred = infer_device_from_id(cid)
                if inferred:
                    parsed["device"] = inferred
            # Use the full id as alarm_name if not set (it IS the firmware alarm identifier)
            if not parsed.get("alarm_name"):
                parsed["alarm_name"] = cid
            entry_map[cid] = parsed
    kb["sources"].append("error_catalog")

    # 3. Pull INV extended descriptions
    print("Pulling INV extended descriptions...")
    inv_paths = [p for p in gh_tree("neurio/pwrinverter")
                 if "ExtendedDescriptions" in p and p.endswith(".md")]
    print(f"  Found {len(inv_paths)} INV extended description files")
    for path in inv_paths:
        alarm_name = Path(path).stem
        content = gh_file_content("neurio/pwrinverter", path)
        if not content:
            continue
        parsed = parse_extended_md(content, alarm_name)
        key = f"INV:{alarm_name}"
        if key in entry_map:
            entry_map[key]["engineering"].update({
                k: v for k, v in parsed["engineering"].items() if v
            })
            entry_map[key]["sources"].append("inv_extended_desc")
        else:
            parsed["device"] = "INV"
            parsed["sources"] = ["inv_extended_desc"]
            entry_map[key] = parsed
    kb["sources"].append("inv_extended_desc")

    # 4. Pull BMU extended descriptions
    print("Pulling BMU extended descriptions...")
    bmu_paths = [p for p in gh_tree("neurio/pwrbmu")
                 if "ExtendedDescriptions" in p and p.endswith(".md")]
    print(f"  Found {len(bmu_paths)} BMU extended description files")
    for path in bmu_paths:
        alarm_name = Path(path).stem
        content = gh_file_content("neurio/pwrbmu", path)
        if not content:
            continue
        parsed = parse_extended_md(content, alarm_name)
        key = f"BMU:{alarm_name}"
        if key in entry_map:
            entry_map[key]["engineering"].update({
                k: v for k, v in parsed["engineering"].items() if v
            })
            entry_map[key]["sources"].append("bmu_extended_desc")
        else:
            parsed["device"] = "BMU"
            parsed["sources"] = ["bmu_extended_desc"]
            entry_map[key] = parsed
    kb["sources"].append("bmu_extended_desc")

    # 5. Pull GMS/Manta headend events from reef
    print("Pulling GMS/Manta headend events...")
    content = gh_file_content("neurio/reef", "source/application/headend/event-manager/headend-events.json")
    if content:
        events = parse_headend_events(content)
        print(f"  Found {len(events)} GMS/Manta alarm events")
        for ev in events:
            key = f"GMS:{ev['alarm_name']}"
            if key in entry_map:
                entry_map[key]["engineering"].update(ev.get("engineering", {}))
                entry_map[key]["sources"].append("reef_headend")
            else:
                ev["sources"] = ["reef_headend"]
                entry_map[key] = ev
    kb["sources"].append("reef_headend")

    # 6. Load supplemental KB from manual/guide content
    print("Loading supplemental knowledge base (manuals, QSGs, guides)...")
    if SUPPLEMENTAL_PATH.exists():
        with open(SUPPLEMENTAL_PATH) as f:
            supp = json.load(f)
        supp_entries = supp.get("entries", [])
        for e in supp_entries:
            key = e.get("id") or f"supplemental:{e.get('alarm_name','')}"
            entry_map[key] = e
        print(f"  {len(supp_entries)} supplemental entries loaded")
        kb["sources"].append("supplemental_manuals")
    else:
        print("  supplemental_kb.json not found — skipping")

    kb["entries"] = list(entry_map.values())
    print(f"\nTotal entries: {len(kb['entries'])}")

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w") as f:
        json.dump(kb, f, indent=2)
    print(f"Written to {OUT_PATH}")


if __name__ == "__main__":
    build()
